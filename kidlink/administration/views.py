from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from .models import Administrator, Youth, YouthActivity, Activity,Institute, YouthInstitute
from .forms import YouthForm, YouthActivityForm, ActivityForm, InstituteForm, YouthInstituteForm
from django.http import HttpResponseForbidden, HttpResponse
# Create your views here.

# administration Authentication
from django.contrib.auth import authenticate, login
from django.shortcuts import render, redirect
from django.contrib import messages

def admin_login(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        user = authenticate(request, username=username, password=password)
        if user is not None and user.is_staff:  # only staff can log in
            login(request, user)
            messages.success(request, f"Welcome back, {user.username}!")
            return redirect("admin_dashboard")
        else:
            return render(request, "administration/login.html", {
                "error": "Invalid credentials or not an administrator."
            })
    return render(request, "administration/login.html")


    
def admin_logout(request):
    logout(request)
    return redirect("admin_login")

@login_required
def admin_dashboard(request):
    from django.db.models import Count
    from datetime import datetime, timedelta
    import json
    
    # Calculate real statistics
    total_youth = Youth.objects.count()
    total_activities = Activity.objects.count()
    total_institutes = Institute.objects.count()
    
    # Calculate monthly participation rate (last 30 days)
    last_30_days = datetime.now() - timedelta(days=30)
    monthly_participations = YouthActivity.objects.filter(
        participation_date__gte=last_30_days
    ).count()
    
    # Get participation trends for the last 12 months
    monthly_data = {}
    for i in range(11, -1, -1):
        month_start = datetime.now().replace(day=1) - timedelta(days=30*i)
        month_end = month_start + timedelta(days=30)
        count = YouthActivity.objects.filter(
            participation_date__gte=month_start.date(),
            participation_date__lt=month_end.date()
        ).count()
        month_label = month_start.strftime('%b')
        monthly_data[month_label] = count
    
    # Get participation by activity category
    activity_data = YouthActivity.objects.values('activity__activity_name').annotate(
        count=Count('id')
    ).order_by('-count')[:5]
    
    activity_names = [item['activity__activity_name'] for item in activity_data] if activity_data else ['No Activities']
    activity_counts = [item['count'] for item in activity_data] if activity_data else [1]
    
    # Recent activities for the sidebar
    recent_activities = YouthActivity.objects.select_related('youth', 'activity').order_by('-participation_date')[:10]
    
    context = {
        'total_youth': total_youth,
        'total_activities': total_activities,
        'total_institutes': total_institutes,
        'monthly_participation_rate': monthly_participations,
        'monthly_labels': json.dumps(list(monthly_data.keys())),
        'monthly_data': json.dumps(list(monthly_data.values())),
        'activity_names': json.dumps(activity_names),
        'activity_counts': json.dumps(activity_counts),
        'recent_activities': recent_activities,
    }
    
    return render(request, "administration/dashboard.html", context)



#funtion for adding activity 
@login_required
def add_activity(request):
    if not request.user.is_staff:  # restrict to admins
        return HttpResponseForbidden("You are not authorized to add activities.")

    if request.method == "POST":
        form = ActivityForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("activity_list")
    else:
        form = ActivityForm()
    return render(request, "activities/add_activity.html", {"form": form})

# funtion to show the activity lists
def activity_list(request):
    activities = Activity.objects.all()
    return render(request, 'activities/activity_list.html', {"activities":activities})

# function to edit an activity
@login_required
def update_activity(request,pk):
    if not request.user.is_staff:
        return HttpResponseForbidden("You are not authorized to edit activities")
    activity = get_object_or_404(Activity, pk=pk)
    if request.method == "POST":
        form = ActivityForm(request.POST, instance=activity)
        if form.is_valid():
            form.save()
            return redirect('activity_list')
    else:
        form = ActivityForm(instance=activity)
        return render(request, 'activities/activity_update.html',{'form':form})
    
# function to delete an acitivty
@login_required
def delete_activity(request, pk):
    if not request.user.is_staff:
        return HttpResponseForbidden("You are not authorized to delete an activity")
    activity = get_object_or_404(Activity, pk=pk)
    if request.method == "POST":
        activity.delete() # deleting  the activity
        return redirect('activity_list') 
    return render(request, 'activities/activity_confirm_delete.html', {'activity': activity})

#functions for insitute
# function to add an institute
@login_required
def add_institute(request):
    if not request.user.is_staff:
        return HttpResponseForbidden("You are not authorized to add an institute")
    if request.method == "POST":
        form = InstituteForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('institute_list')
    else:
        form = InstituteForm()
    return render(request, "institutes/add_institute.html", {"form": form})

@login_required   
# showing lists of institutes
def institute_list(request):
    if not request.user.is_staff:
        return HttpResponseForbidden("You are not authorized to add an institute")
    institutes = Institute.objects.all()
    return render(request, 'institutes/institute_list.html', {"institutes":institutes})

# updating an institute
@login_required
def update_institute(request,pk):
    if not request.user.is_staff:
        return HttpResponseForbidden("You are not authorized to edit an institute")
    institute = get_object_or_404(Institute, pk=pk)
    if request.method == "POST":
        form = InstituteForm(request.POST, instance=institute)
        if form.is_valid():
            form.save()
            return redirect('institute_list')
    else:
        form = InstituteForm(instance=institute)
    return render(request, 'institutes/institute_update.html',{'form':form})
    
# function to delete an institute
@login_required
def delete_institute(request, pk):
    if not request.user.is_staff:
        return HttpResponseForbidden("You are not authorized to delete an activity")
    institute = get_object_or_404(Institute, pk=pk)
    if request.method == "POST":
        institute.delete() # deleting  the activity
        return redirect('institute_list') 
    return render(request, 'institutes/institute_confirm_delete.html', {'institute': institute})

# Youth Activity Administration
@login_required
def youth_activity_list(request):
    if not request.user.is_staff:
        return HttpResponseForbidden("You are not authorized to view this page.")
    activities = YouthActivity.objects.all()
    return render(request, 'administration/youth_activity_list.html', {'activities': activities})

@login_required
def youth_activity_create(request):
    if not request.user.is_staff:
        return HttpResponseForbidden("You are not authorized to add activities.")
    if request.method == "POST":
        form = YouthActivityForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('youth_activity_list')
    else:
        form = YouthActivityForm()
    return render(request, 'administration/youth_activity_create.html', {'form': form})

@login_required
def youth_activity_update(request, pk):
    if not request.user.is_staff:
        return HttpResponseForbidden("You are not authorized to edit activities.")
    activity = get_object_or_404(YouthActivity, pk=pk)
    if request.method == "POST":
        form = YouthActivityForm(request.POST, instance=activity)
        if form.is_valid():
            form.save()
            return redirect('youth_activity_list')
    else:
        form = YouthActivityForm(instance=activity)
    return render(request, 'administration/youth_activity_form.html', {'form': form})

@login_required
def youth_activity_delete(request, pk):
    if not request.user.is_staff:
        return HttpResponseForbidden("You are not authorized to delete activities.")
    activity = get_object_or_404(YouthActivity, pk=pk)
    if request.method == "POST":
        activity.delete()
        return redirect('youth_activity_list')
    return render(request, 'administration/youth_activity_confirm_delete.html', {'activity': activity})


# LOLLLLLLLLLLL
# Youth Administration
@login_required
def youth_list(request):
    if not request.user.is_staff:
        return HttpResponseForbidden("You are not authorized to view this page.")
    youths = Youth.objects.all()
    return render(request, 'administration/youth_list.html', {'youths': youths})

# for creating Youth
@login_required
def youth_create(request):
    if not request.user.is_staff:  # restrict to admins
        return HttpResponseForbidden("You are not authorized to add youths.")
    if request.method == "POST":
        form = YouthForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('youth_list')
    else:
        form = YouthForm()
    return render(request, 'administration/youth_form.html', {'form': form})

# for updating youth
@login_required
def youth_update(request, pk):
    if not request.user.is_staff:
        return HttpResponseForbidden("You are not authorized to edit youths.")
    youth = get_object_or_404(Youth, pk=pk)
    if request.method == "POST":
        form = YouthForm(request.POST, instance=youth)
        if form.is_valid():
            form.save()
            return redirect('youth_list')
    else:
        form = YouthForm(instance=youth)
    return render(request,'administration/youth_form.html', {'form': form})

# delete youth
@login_required
def youth_delete(request, pk):
    if not request.user.is_staff:
        return HttpResponseForbidden("You are not authorized to delete youths.")
    youth = get_object_or_404(Youth, pk=pk)
    if request.method == "POST":
        youth.delete()
        return redirect('youth_list')
    return render(request, 'administration/youth_confirm_delete.html', {'youth': youth})

# Youth institute Administration
@login_required
def youth_institute_list(request):
    youth_institutes = YouthInstitute.objects.all()
    return render(request, 'youthInstitute/youth_institute_list.html', {'youth_institutes': youth_institutes})

@login_required
def youth_institute_create(request):
    if not request.user.is_staff:
        return HttpResponseForbidden("You are not authorized to add institutes.")
    if request.method == "POST":
        form = YouthInstituteForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('youth_institute_list')
    else:
        form = YouthInstituteForm()
    return render(request, 'youthInstitute/youth_institute_create.html', {'form': form})

@login_required
def youth_institute_update(request, pk):
    if not request.user.is_staff:
        return HttpResponseForbidden("You are not authorized to edit institutes.")
    institute = get_object_or_404(YouthInstitute, pk=pk)
    if request.method == "POST":
        form = YouthInstituteForm(request.POST, instance=institute)
        if form.is_valid():
            form.save()
            return redirect('youth_institute_list')
    else:
        form = YouthInstituteForm(instance=institute)
    return render(request, 'youthInstitute/youth_institute_form.html', {'form': form})

@login_required
def youth_institute_delete(request, pk):
    if not request.user.is_staff:
        return HttpResponseForbidden("You are not authorized to delete institutes.")
    institute = get_object_or_404(YouthInstitute, pk=pk)
    if request.method == "POST":
        institute.delete()
        return redirect('youth_institute_list')
    return render(request, 'youthInstitute/youth_institute_confirm_delete.html', {'institute': institute})

# Export Functions
@login_required
def export_youth_excel(request):
    """Export youth data to Excel format"""
    if not request.user.is_staff:
        return HttpResponseForbidden("You are not authorized to export data.")
    
    from openpyxl import Workbook
    from django.http import HttpResponse
    from datetime import datetime
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Youth Data"
    
    # Add headers
    headers = ['First Name', 'Last Name', 'Date of Birth', 'Gender', 'Status', 'Registration Due']
    ws.append(headers)
    
    # Style headers
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill(start_color="85BA49", end_color="85BA49", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Add data
    youths = Youth.objects.all()
    for youth in youths:
        ws.append([
            youth.first_name,
            youth.last_name,
            youth.date_of_birth.strftime('%Y-%m-%d'),
            youth.gender,
            youth.status,
            youth.registration_due.strftime('%Y-%m-%d')
        ])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="youth_data_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response

@login_required
def export_activities_pdf(request):
    """Export activities data to PDF format"""
    if not request.user.is_staff:
        return HttpResponseForbidden("You are not authorized to export data.")
    
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch
    from django.http import HttpResponse
    from datetime import datetime
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="activities_report_{datetime.now().strftime("%Y%m%d")}.pdf"'
    
    # Create PDF
    doc = SimpleDocTemplate(response, pagesize=A4)
    story = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = styles['Title']
    title_style.textColor = colors.HexColor('#85BA49')
    
    # Title
    title = Paragraph("KidLink Activities Report", title_style)
    story.append(title)
    story.append(Spacer(1, 0.2*inch))
    
    # Activity data
    activities = Activity.objects.all()
    
    # Prepare table data
    data = [['Activity Name', 'Description', 'Start Date', 'End Date', 'Location']]
    
    for activity in activities:
        desc = activity.description[:50] + '...' if len(activity.description) > 50 else activity.description
        data.append([
            activity.activity_name,
            desc,
            activity.start_date.strftime('%Y-%m-%d'),
            activity.end_date.strftime('%Y-%m-%d'),
            activity.location
        ])
    
    # Create table
    table = Table(data, colWidths=[2*inch, 2*inch, 1*inch, 1*inch, 1.5*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#85BA49')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(table)
    
    # Build PDF
    doc.build(story)
    return response

@login_required
def export_full_report(request):
    """Export full comprehensive report"""
    if not request.user.is_staff:
        return HttpResponseForbidden("You are not authorized to export data.")
    
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch
    from django.http import HttpResponse
    from datetime import datetime
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="full_report_{datetime.now().strftime("%Y%m%d")}.pdf"'
    
    # Create PDF
    doc = SimpleDocTemplate(response, pagesize=A4)
    story = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = styles['Title']
    title_style.textColor = colors.HexColor('#85BA49')
    heading_style = styles['Heading1']
    
    # Title
    title = Paragraph("KidLink Comprehensive Report", title_style)
    story.append(title)
    story.append(Spacer(1, 0.3*inch))
    
    # Statistics
    total_youth = Youth.objects.count()
    total_activities = Activity.objects.count()
    total_institutes = Institute.objects.count()
    total_participations = YouthActivity.objects.count()
    
    stats_data = [
        ['Metric', 'Count'],
        ['Total Youth Enrolled', str(total_youth)],
        ['Total Activities', str(total_activities)],
        ['Total Institutes', str(total_institutes)],
        ['Total Participations', str(total_participations)]
    ]
    
    stats_table = Table(stats_data, colWidths=[3*inch, 1*inch])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#85BA49')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(Paragraph("Summary Statistics", heading_style))
    story.append(Spacer(1, 0.1*inch))
    story.append(stats_table)
    story.append(PageBreak())
    
    # Youth List
    youths = Youth.objects.all()[:20]  # Limit to first 20 for PDF
    youth_data = [['First Name', 'Last Name', 'Date of Birth', 'Gender', 'Status']]
    
    for youth in youths:
        youth_data.append([
            youth.first_name,
            youth.last_name,
            youth.date_of_birth.strftime('%Y-%m-%d'),
            youth.gender,
            youth.status
        ])
    
    youth_table = Table(youth_data, colWidths=[1.2*inch, 1.2*inch, 1*inch, 0.8*inch, 1.2*inch])
    youth_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#85BA49')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8)
    ]))
    
    story.append(Paragraph("Youth List", heading_style))
    story.append(Spacer(1, 0.1*inch))
    story.append(youth_table)
    
    # Build PDF
    doc.build(story)
    return response